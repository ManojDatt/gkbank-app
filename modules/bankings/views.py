from django.views.generic import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
import openpyxl, pdb
from openpyxl.styles import Font, Alignment
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from datetime import datetime
from modules.bankings.models import TransactionReport

class DownloadBankStatement(LoginRequiredMixin, View):
	def post(self, request, *args, **kwargs):
		now = datetime.now()
		statement_title = f"Bank-statement-{int(datetime.timestamp(now))}"
		querysets = TransactionReport.objects.filter(translation__customer_id__in=request.POST.get('customers').split(','))
		response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = f'attachment; filename={statement_title}.xlsx'
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = statement_title
		row_num = 0
		columns = [
			(u"Label", 30),
			(u"A/c. Number", 30),
			(u"Action", 20),
			(u"Deposit", 20),
			(u"Withdraw", 20),
			(u"Total", 20),
			(u"Date time", 20),
		]

		for col_num in range(len(columns)):
			cl = ws.cell(row=row_num + 1, column=col_num + 1)
			cl.value = columns[col_num][0]
			cl.font = Font(size=14, bold=True)
			# set column width
			ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
		for obj in querysets:
			row_num += 1
			row = [
				obj.label,
				obj.ac_number,
				obj.trans_type,
				obj.amount_deposit,
				obj.amount_withdraw,
				obj.amount_total,
				obj.updated_at
			]
			for col_num in range(len(row)):
				cl = ws.cell(row=row_num + 1, column=col_num + 1)
				cl.value = row[col_num]
				cl.alignment = Alignment(wrap_text=True)

		wb.save(response)
		return response